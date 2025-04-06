import os
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ARCHIVO_MAESTRO = 'maestro.xlsx'
CARPETA_LISTAS = 'Excels'

# Función para encontrar la columna que contiene "Telefono"
def encontrar_columna_telefono(df):
    for i, row in df.iterrows():
        for col in df.columns:
            valor = str(row[col]).strip().lower()
            if valor == 'telefono':
                return i, col
    return None, None

# --------------------
# Modo RESET si se pasa argumento
# --------------------
if len(sys.argv) > 1 and sys.argv[1].lower() == 'reset':
    df_vacio = pd.DataFrame([['Telefono']])
    df_vacio.to_excel(ARCHIVO_MAESTRO, index=False, header=False)
    print("🔁 Maestro reseteado. Archivos anteriores ignorados.")
    sys.exit()

# --------------------
# Cargar datos del maestro
# --------------------
try:
    df_maestro_raw = pd.read_excel(ARCHIVO_MAESTRO, header=None)
    fila_titulo_maestro, col_telefono_maestro = encontrar_columna_telefono(df_maestro_raw)

    if fila_titulo_maestro is None:
        print(f"⚠️ No se encontró columna 'Telefono' en {ARCHIVO_MAESTRO}")
        telefonos_existentes = []
    else:
        idx_telefono_maestro = df_maestro_raw.columns.get_loc(col_telefono_maestro)
        telefonos_existentes = df_maestro_raw.iloc[fila_titulo_maestro+1:, idx_telefono_maestro].dropna()
        telefonos_existentes = telefonos_existentes.astype(str).str.strip().tolist()

    print(f"📖 Maestro cargado con {len(telefonos_existentes)} registros.")
except FileNotFoundError:
    print(f"⚠️ Archivo {ARCHIVO_MAESTRO} no encontrado. Se creará uno nuevo.")
    telefonos_existentes = []

# --------------------
# Leer archivos desde carpeta "Excels"
# --------------------
datos_nuevos = []

if not os.path.exists(CARPETA_LISTAS):
    print(f"📁 Carpeta '{CARPETA_LISTAS}' no encontrada.")
else:
    archivos = [f for f in os.listdir(CARPETA_LISTAS) if f.endswith('.xlsx')]
    print(f"🗂 Archivos encontrados: {archivos}")

    for archivo in archivos:
        ruta_archivo = os.path.join(CARPETA_LISTAS, archivo)
        print(f"🔍 Leyendo: {archivo}")
        try:
            df = pd.read_excel(ruta_archivo, header=None)

            fila_titulo, col_telefono = encontrar_columna_telefono(df)
            if fila_titulo is None:
                print(f"⚠️ No se encontró columna 'Telefono' en {archivo}")
                continue

            col_idx = df.columns.get_loc(col_telefono)
            telefonos = df.iloc[fila_titulo+1:, col_idx].dropna()
            telefonos = telefonos.astype(str).str.strip()

            datos_nuevos.extend(telefonos.tolist())

        except Exception as e:
            print(f"❌ Error leyendo {archivo}: {e}")

# --------------------
# Comparar y guardar en ubicación exacta
# --------------------
telefonos_para_agregar = [t for t in datos_nuevos if t not in telefonos_existentes]

if telefonos_para_agregar:
    # Abrir workbook y hoja activa
    wb = load_workbook(ARCHIVO_MAESTRO)
    ws = wb.active

    # Buscar coordenadas de 'Telefono'
    fila_titulo, col_letra = None, None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if str(cell.value).strip().lower() == 'telefono':
                fila_titulo = cell.row
                col_letra = get_column_letter(cell.column)
                break
        if fila_titulo:
            break

    if not fila_titulo or not col_letra:
        print("⚠️ No se encontró la celda con título 'Telefono'.")
    else:
        # Buscar primera fila vacía debajo del título
        fila_actual = fila_titulo + 1
        while ws[f"{col_letra}{fila_actual}"].value is not None:
            fila_actual += 1

        # Insertar nuevos teléfonos
        for tel in telefonos_para_agregar:
            ws[f"{col_letra}{fila_actual}"] = tel
            fila_actual += 1

        wb.save(ARCHIVO_MAESTRO)
        print(f"✅ {len(telefonos_para_agregar)} teléfonos agregados al maestro en columna {col_letra} debajo del título.")
else:
    print("ℹ️ No se encontraron datos nuevos para agregar.")
